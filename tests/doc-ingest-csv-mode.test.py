#!/usr/bin/env python3
"""Tests for doc-ingest's --csv compute-exact mode (child of the doc-ingest skill).

The reading view (markdown, row-capped, summarized) answers "what does this file
say"; --csv answers "give me the exact table so I can COMPUTE" — the GAIA-validated
rule that quantitative answers must come from computation over the loaded table
(3/3 computable misses flipped, 84.2%→92.1%, 2026-07-30). Pinned behaviors:

  1. csv/tsv → exact CSV out (tsv normalized), no row cap, values verbatim
  2. --csv output is NOT truncated by the DEFAULT char cap (explicit --max-chars still wins)
  3. xlsx → per-sheet '=== sheet: NAME ===' CSV via openpyxl (skipped if not importable)
  4. xlsx WITHOUT openpyxl → clear refusal (never the approximate zip fallback)
  5. non-tabular file with --csv → per-file error, exit 1
  6. computing over the emitted CSV reproduces an exact aggregate (the point)

Run: python3 tests/doc-ingest-csv-mode.test.py   (exit 0 pass / 1 fail)
"""
from __future__ import annotations

import builtins
import csv
import importlib.util
import io
import json
import sys
import tempfile
import types
import zipfile
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
spec = importlib.util.spec_from_file_location(
    "doc_ingest", REPO / "skills" / "doc-ingest" / "scripts" / "ingest.py"
)
ingest = importlib.util.module_from_spec(spec)
spec.loader.exec_module(ingest)

failures: list[str] = []


def check(name: str, cond: bool, detail: str = "") -> None:
    print(("ok   " if cond else "FAIL ") + name + ("" if cond else f" — {detail}"))
    if not cond:
        failures.append(name)


def run_main(argv):
    out, err = io.StringIO(), io.StringIO()
    with redirect_stdout(out), redirect_stderr(err):
        rc = ingest.main(argv)
    return rc, out.getvalue(), err.getvalue()


tmp = Path(tempfile.mkdtemp(prefix="doc-ingest-csv-"))

# 1. csv → exact CSV, no row cap, verbatim values
big = tmp / "big.csv"
with open(big, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(["name", "wheels"])
    for i in range(600):  # beyond the 500-row reading cap
        w.writerow([f"loco{i}", i % 7])
rc, out, err = run_main([str(big), "--csv"])
lines = [ln for ln in out.strip().splitlines() if ln]
check("csv: exit 0", rc == 0, f"rc={rc} err={err}")
check("csv: all 601 rows present (no reading cap)", len(lines) == 601, f"lines={len(lines)}")
check("csv: values verbatim", lines[0] == "name,wheels" and lines[1] == "loco0,0", repr(lines[:2]))

# 1b. tsv normalized to csv
tsv = tmp / "t.tsv"
tsv.write_text("a\tb\n1\t2\n")
rc, out, _ = run_main([str(tsv), "--csv"])
check("tsv: normalized to csv", rc == 0 and out.strip().splitlines() == ["a,b", "1,2"], repr(out))

# 2. default char cap does not truncate --csv; explicit --max-chars still wins
wide = tmp / "wide.csv"
with open(wide, "w", newline="") as fh:
    w = csv.writer(fh)
    for i in range(3000):
        w.writerow([f"row{i}", "x" * 90])  # ~ 280k chars total > 200k default cap
rc, out, _ = run_main([str(wide), "--csv"])
check("csv: default cap lifted (no truncation notice)", rc == 0 and "truncated" not in out,
      f"len={len(out)}")
check("csv: full content beyond 200k chars", len(out) > 250_000, f"len={len(out)}")
rc, out, _ = run_main([str(wide), "--csv", "--max-chars", "1000"])
check("csv: explicit --max-chars still wins", len(out) < 2000 and "truncated" in out, f"len={len(out)}")

# 3+6. xlsx per-sheet CSV + exact aggregate (needs openpyxl; skip gracefully)
try:
    import openpyxl  # noqa: F401
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False
if HAVE_OPENPYXL:
    from openpyxl import Workbook
    xlsx = tmp / "wheels.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Steam"
    ws.append(["config", "wheels"])
    for cfg, n in (("0-4-0", 4), ("2-8-4", 14), ("4-4-0", 8)):
        ws.append([cfg, n])
    ws2 = wb.create_sheet("Diesel")
    ws2.append(["unit", "hp"])
    ws2.append(["GP7", 1500])
    wb.save(xlsx)
    rc, out, _ = run_main([str(xlsx), "--csv"])
    check("xlsx: exit 0", rc == 0)
    check("xlsx: per-sheet markers", "=== sheet: Steam ===" in out and "=== sheet: Diesel ===" in out, out[:120])
    # 6. the point: computing over the emitted CSV gives the exact aggregate
    steam_block = out.split("=== sheet: Steam ===")[1].split("=== sheet:")[0].strip()
    rows = list(csv.reader(io.StringIO(steam_block)))
    total = sum(int(r[1]) for r in rows[1:])
    check("xlsx: exact aggregate from emitted CSV (4+14+8=26)", total == 26, f"total={total}")
else:
    print("skip xlsx cases (openpyxl not importable) — refusal case below still runs")

# 3b. Host-independent coverage of the openpyxl load_workbook path (ingest.py
# 788-793). The happy-path test above SKIPS when openpyxl is not importable —
# which is exactly the hosted diff-coverage env — leaving 788-793 uncovered
# (qingyun CR #2434). Inject a stub openpyxl so the load_workbook -> per-sheet
# iteration path executes regardless of whether the real dep is installed.
stub_xlsx = tmp / "stub.xlsx"
with zipfile.ZipFile(stub_xlsx, "w", zipfile.ZIP_DEFLATED) as zf:
    zf.writestr("xl/workbook.xml", b"<x/>")  # valid small zip, uncompressed well under budget


class _StubWS:
    def __init__(self, title, rows):
        self.title = title
        self._rows = rows

    def iter_rows(self, values_only=True):  # noqa: ARG002
        return iter(self._rows)


class _StubWB:
    def __init__(self, sheets):
        self.worksheets = sheets


def _stub_load_workbook(path, read_only=True, data_only=True):  # noqa: ARG001
    return _StubWB(
        [
            _StubWS("Steam", [("config", "wheels"), ("0-4-0", 4), ("2-8-4", 14), ("4-4-0", 8)]),
            _StubWS("Diesel", [("unit", "hp"), ("GP7", 1500)]),
        ]
    )


stub_openpyxl = types.ModuleType("openpyxl")
stub_openpyxl.load_workbook = _stub_load_workbook  # type: ignore[attr-defined]
_saved_openpyxl = sys.modules.get("openpyxl")
sys.modules["openpyxl"] = stub_openpyxl
try:
    stub_out = ingest.extract_table_csv(stub_xlsx)
    check(
        "xlsx (stub openpyxl): per-sheet markers — covers load_workbook path",
        "=== sheet: Steam ===" in stub_out and "=== sheet: Diesel ===" in stub_out,
        stub_out[:160],
    )
    _steam = stub_out.split("=== sheet: Steam ===")[1].split("=== sheet:")[0].strip()
    _rows = list(csv.reader(io.StringIO(_steam)))
    check(
        "xlsx (stub openpyxl): exact aggregate over emitted CSV (4+14+8=26)",
        sum(int(r[1]) for r in _rows[1:]) == 26,
        _steam,
    )
finally:
    if _saved_openpyxl is not None:
        sys.modules["openpyxl"] = _saved_openpyxl
    else:
        sys.modules.pop("openpyxl", None)

# 4. xlsx without openpyxl → clear refusal, never the approximate fallback
xlsx_path = tmp / "any.xlsx"
xlsx_path.write_bytes(b"PK\x03\x04fake")
real_import = builtins.__import__


def _no_openpyxl(name, *a, **k):
    if name == "openpyxl":
        raise ImportError("blocked for test")
    return real_import(name, *a, **k)


builtins.__import__ = _no_openpyxl
try:
    try:
        ingest.extract_table_csv(xlsx_path)
        check("xlsx no-openpyxl: refused", False, "no exception raised")
    except RuntimeError as e:
        check("xlsx no-openpyxl: refused with clear error", "openpyxl" in str(e), str(e))
    except Exception as e:  # noqa: BLE001
        check("xlsx no-openpyxl: refused", False, f"wrong exception {type(e).__name__}: {e}")
finally:
    builtins.__import__ = real_import

# 5. non-tabular with --csv → per-file error, exit 1; json carries the error
txt = tmp / "notes.txt"
txt.write_text("hello")
rc, out, err = run_main([str(txt), "--csv", "--json"])
check("non-tabular: exit 1", rc == 1, f"rc={rc}")
rec = json.loads(out.strip().splitlines()[0])
check("non-tabular: json error names the constraint", not rec["ok"] and "--csv" in rec["error"], repr(rec))

# 6. resource budgets are fail-closed (never truncate, never OOM): an input over
# the source-bytes budget is REFUSED with a loud error naming the override, and
# the explicit --csv-no-budget override restores the exact full output.
budget_csv = tmp / "budget.csv"
with open(budget_csv, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(["n", "v"])
    for i in range(50):
        w.writerow([i, i * 3])
real_src_budget = ingest.CSV_MAX_SOURCE_BYTES
ingest.CSV_MAX_SOURCE_BYTES = 64  # smaller than the file
try:
    rc, out, err = run_main([str(budget_csv), "--csv"])
    check("oversized source: refused, exit 1",
          rc == 1 and "compute budget" in err and "--csv-no-budget" in err,
          f"rc={rc} err={err.strip()[:160]}")
    check("oversized source: nothing on stdout (no partial table)", out == "", repr(out[:80]))
    rc, out, err = run_main([str(budget_csv), "--csv", "--csv-no-budget"])
    check("oversized source + --csv-no-budget: exact full output",
          rc == 0 and out.strip().splitlines()[-1] == "49,147",
          f"rc={rc} tail={out.strip().splitlines()[-1] if out.strip() else '<empty>'}")
finally:
    ingest.CSV_MAX_SOURCE_BYTES = real_src_budget

# 6b. OOXML zip-bomb: a SMALL compressed archive whose UNCOMPRESSED payload is
# over the source budget must be refused BEFORE openpyxl parses it — the
# compressed-size gate alone can't see the inflation (qingyun CR #2434). Proven
# host-independently by blocking the openpyxl import and asserting the inflation
# error fires anyway, i.e. load_workbook is never reached.
bomb = tmp / "bomb.xlsx"
with zipfile.ZipFile(bomb, "w", zipfile.ZIP_DEFLATED) as zf:
    zf.writestr("xl/sharedStrings.xml", b"A" * 4096)  # ~tiny compressed, 4096 uncompressed
real_src_budget2 = ingest.CSV_MAX_SOURCE_BYTES
ingest.CSV_MAX_SOURCE_BYTES = 1024  # compressed file < 1024 (passes), uncompressed 4096 > 1024
try:
    check("zip-bomb: compressed file is under the source budget",
          bomb.stat().st_size < 1024, f"size={bomb.stat().st_size}")
    builtins.__import__ = _no_openpyxl  # if the preflight fires, openpyxl is never imported
    try:
        try:
            ingest.extract_table_csv(bomb)
            check("zip-bomb: refused before openpyxl", False, "no exception raised")
        except RuntimeError as e:
            check("zip-bomb: refused on UNCOMPRESSED size before openpyxl",
                  "uncompressed" in str(e) and "openpyxl" not in str(e), str(e)[:160])
        except Exception as e:  # noqa: BLE001
            check("zip-bomb: refused before openpyxl", False,
                  f"wrong exception {type(e).__name__}: {e}")
    finally:
        builtins.__import__ = real_import
    # --csv-no-budget opts out of the inflation preflight: it should reach the
    # openpyxl path (here blocked → the openpyxl-needed error, NOT the inflation error).
    builtins.__import__ = _no_openpyxl
    try:
        try:
            ingest.extract_table_csv(bomb, unbounded=True)
            check("zip-bomb + --csv-no-budget: preflight skipped", False, "no exception raised")
        except RuntimeError as e:
            check("zip-bomb + --csv-no-budget: skips inflation preflight (reaches openpyxl)",
                  "openpyxl" in str(e), str(e)[:160])
    finally:
        builtins.__import__ = real_import
finally:
    ingest.CSV_MAX_SOURCE_BYTES = real_src_budget2

# 7. render budgets bound cells and output bytes even when the file-size gate
# passes — and the error is loud, not a truncation.
real_cells = ingest.MAX_TABLE_CELLS
ingest.MAX_TABLE_CELLS = 10  # 51 rows x 2 cells blows this immediately
try:
    rc, out, err = run_main([str(budget_csv), "--csv"])
    check("cell budget: refused loudly",
          rc == 1 and "cell compute" in err and "--csv-no-budget" in err,
          f"rc={rc} err={err.strip()[:160]}")
finally:
    ingest.MAX_TABLE_CELLS = real_cells
real_out_budget = ingest.CSV_MAX_OUTPUT_BYTES
ingest.CSV_MAX_OUTPUT_BYTES = 40  # a few rows of output blow this
try:
    rc, out, err = run_main([str(budget_csv), "--csv"])
    check("output-bytes budget: refused loudly",
          rc == 1 and "compute" in err and "--csv-no-budget" in err,
          f"rc={rc} err={err.strip()[:160]}")
finally:
    ingest.CSV_MAX_OUTPUT_BYTES = real_out_budget
rc, out, err = run_main([str(budget_csv), "--csv"])
check("budgets restored: normal-sized table still exact",
      rc == 0 and out.strip().splitlines()[-1] == "49,147", f"rc={rc}")

# 7b. the output-byte budget counts ENCODED UTF-8 bytes, not characters
# (qingyun CR 2026-07-31): replacement characters (U+FFFD, 3 UTF-8 bytes each)
# and other multibyte content must charge their rendered byte cost — otherwise
# an untrusted non-ASCII attachment bypasses CSV_MAX_OUTPUT_BYTES while --csv
# exits 0. Reproduces the review probe: a ~30-byte invalid-UTF-8 row renders to
# ~32 characters but ~92 encoded bytes.
mb_csv = tmp / "multibyte.csv"
mb_csv.write_bytes(b"\xff" * 30 + b"\n")  # errors="replace" -> 30 U+FFFD chars
real_out_budget = ingest.CSV_MAX_OUTPUT_BYTES
ingest.CSV_MAX_OUTPUT_BYTES = 40  # chars fit (32), encoded bytes do not (92)
try:
    rc, out, err = run_main([str(mb_csv), "--csv"])
    check("byte budget counts encoded bytes (replacement chars): refused",
          rc == 1 and "compute" in err and "--csv-no-budget" in err,
          f"rc={rc} err={err.strip()[:160]}")
    check("byte budget (replacement chars): nothing on stdout (no partial table)",
          out == "", repr(out[:80]))
finally:
    ingest.CSV_MAX_OUTPUT_BYTES = real_out_budget

euro_csv = tmp / "euro.csv"
euro_csv.write_text("€" * 20 + "\n", encoding="utf-8")  # 22 chars/row, 62 bytes
real_out_budget = ingest.CSV_MAX_OUTPUT_BYTES
ingest.CSV_MAX_OUTPUT_BYTES = 40  # would pass on a character count
try:
    rc, out, err = run_main([str(euro_csv), "--csv"])
    check("byte budget counts encoded bytes (multibyte): refused",
          rc == 1 and "compute" in err, f"rc={rc} err={err.strip()[:160]}")
    ingest.CSV_MAX_OUTPUT_BYTES = 128  # encoded bytes genuinely fit
    rc, out, err = run_main([str(euro_csv), "--csv"])
    check("multibyte under the byte budget: exact output, exit 0",
          rc == 0 and out.strip() == "€" * 20, f"rc={rc} out={out[:40]!r}")
finally:
    ingest.CSV_MAX_OUTPUT_BYTES = real_out_budget

print()
if failures:
    print(f"{len(failures)} FAILURE(S): {failures}")
    sys.exit(1)
print("All doc-ingest --csv checks passed.")
